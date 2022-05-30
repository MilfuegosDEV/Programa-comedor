# Módulos
from tkinter import END, Entry, PhotoImage, Tk, Label, Button, Toplevel, messagebox
from tkinter.filedialog import askopenfilenames; 
from PIL import Image, ImageTk
import pandas as pd 
from openpyxl import load_workbook, Workbook 
from datetime import datetime
import time 
import shutil
import sys
import os 
import json


PersonasDelComedor = {} # guardará los datos de las personas que estan en el archivo
UsaronElComedor = {} # Datos de todas las personas que utilizaron el comedor.
CedulasRegistradas = [] # cedulas que fueron registradas durante la ejecución del programa.

actual = time.strftime('%d-%m-%y') 
Font_tuple = ("Comic Sans MS", 20)
file_tuples = ('Excel files', '.xlsx'), ('All files', '.')






def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)








def cargar_json():
    """Se cargar los datos de las personas que fueron al comedor durante el día"""
    global CedulasRegistradas

    try:
        with open(f'C:\SistemaComedor\Diario\RegistroDelDía{actual}.json', 'r') as File:
            CedulasRegistradas = json.load(File)
    except FileNotFoundError:
        editar_json()





def editar_json():
    """Edita el reporte diario, o directamente lo crea"""
    os.makedirs('C:\SistemaComedor\Diario', exist_ok= True) 
    os.system('attrib +h C:\SistemaComedor\Diario')
    
    with open(f'C:\SistemaComedor\Diario\RegistroDelDía{actual}.json', 'w') as File:
        json.dump(CedulasRegistradas, File, indent=4)
        File.close()





def isnotEmpty(data_structure): 
    """Verifica los archivos no esten vacíos"""
    if (len(data_structure) > 0):
        # si el archivo no esta vacio retorna true
        return True

    else:
        # retorna false
        messagebox.showwarning('FORMATO INCORRECTO', 'EL ARCHIVO DEBE IR EN EL SIGUIENTE FORMATO\nCedula, Nombre, Sección, Grupo')
        os.startfile(r'C:\SistemaComedor\Comedor.xlsx')
        return False







def Archivo_no_encontrado(Archivo): 
    """Si el archivo no esta el programa"""
    messagebox.showerror('ERROR ARCHIVO NO ENCONTRADO', f'PRESIONE ABRIR PARA MOVER EL ARCHIVO {Archivo} al diretorio C:\SistemaComedor')
    source = askopenfilenames(title = 'Mover',filetypes=(file_tuples))
    destination = 'C:\SistemaComedor'

    for files in source:
        try:
            shutil.move(files, destination)
        except shutil.Error:
            pass
    return False




def VERIFICACION_ARCHIVO(Archivo):
    """La siguiente función verifica la integridad y el formato de los datos que se encuentran en 
    en el archivo Comedor.xlsx"""
    datos = [] 
    info = [] 
    
    try:
        # verifica si el archivo esta en la carpeta y también el formato de los datos.
        df = pd.read_excel(r'C:\SistemaComedor\Comedor.xlsx') 
        for values in df.itertuples(index= False, name= None):
            datos.append(values)

        for cedula, nombre, seccion, grupo in datos:
            if str(cedula).isalnum() == True:
                info.append(nombre); info.append(seccion); info.append(grupo)
                PersonasDelComedor[str(cedula).upper()] = tuple(info)
                info.clear()

            elif str(cedula).isnumeric():
                info.append(nombre); info.append(seccion); info.append(grupo)
                PersonasDelComedor[str(cedula)] = tuple(info)
                info.clear()

        datos.clear() 

        for clave, valor0 in PersonasDelComedor.items():
            # Verifica si el lugar donde están deberían estar los números de cédulas
            # esta en el formato correcto.
            if str(clave).isalpha() == True or str(valor0[0]).isalpha() == False:
                PersonasDelComedor.clear()
                break

        if isnotEmpty(PersonasDelComedor) == True:
            return True
        
    except FileNotFoundError:
        Archivo_no_encontrado(Archivo)

    except ValueError:
        messagebox.showwarning('FORMATO INCORRECTO', 'EL ARCHIVO DEBE IR EN EL SIGUIENTE FORMATO\nCedula, Nombre, Sección, Grupo')
        os.startfiles(r'C:\SistemaComedor\Comedor.xlsx')
        return False








def GuardarRegistro():
    """Guarda el registro de las personas que utilizaron el comedor."""
    global UsaronElComedor
    
    fecha = str(datetime.today().strftime('%m-%y')) # mes-año
    hoy = pd.to_datetime(datetime.today().strftime('%d/%m/%y %H:%M:%S'), dayfirst= True) # dia-mes-año hora-minutos-segundos 

    os.makedirs(r'C:\SistemaComedor\reportes', exist_ok=True) # crea el directorio donde se guardarán los reportes.

    while True:
        try:
            archivo = load_workbook(r'C:\SistemaComedor\reportes\Reporte '+ fecha+'.xlsx')
            ws = archivo.active # worksheet 
            i = ws.max_row; i += 1 # busca la ultima celda sin datos.

            for cedula, datos in UsaronElComedor.items():
                # si la persona esta en la base de datos la guarda con sus datos los cuales son 
                #| Numero de cédula | Nombre | Sección |  Grupo | Fecha
                ws[f'A{i}'] = cedula # cedula 
                ws[f'B{i}'] = datos[0] # Nombre
                ws[f'C{i}'] = datos[1] # Seccion
                ws[f'D{i}'] = datos[2] # Grupo
                ws[f'E{i}'] = hoy # fecha
                i += 1 # avanza una linea.

            archivo.save(r'C:\SistemaComedor\reportes\Reporte '+ fecha+'.xlsx')
            break

        except FileNotFoundError:
            # si no esta el archivo lo genera y le introduce los encabezados
            wb = Workbook()
            ws = wb.active
            ws.title = 'Registro'
            ws.append({1:'\tCEDULA', 2:'\tNOMBRE', 3: "\tSECCIÓN", 4:'\tGRUPO', 5:'\tFECHA'})
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            wb.save(fr'C:\SistemaComedor\reportes\Reporte '+fecha+'.xlsx')
            continue

        except PermissionError:
            # Solo si el archivo de reportes esta abierto.
            messagebox.showwarning('Reporte abierto',f"Por Favor cierre el archivo antes de continuar 'Reporte {fecha}.xlsx'")
            continue

        except ValueError:
            # en
            messagebox.showwarning('FORMATO INCORRECTO', 'EL ARCHIVO DEBE IR EN EL SIGUIENTE FORMATO\nCedula, Nombre, Sección, Grupo')
            os.startfiles(r'C:\SistemaComedor\Comedor.xlsx')
            return False






def abrir_reportes():
    """Abre un menu para seleccionar los reportes que se quieren abrir"""
    source = askopenfilenames(initialdir= r'C:\SistemaComedor\Reportes', filetypes=(file_tuples)) # se toma el directorio donde se 
    for files in source:
        os.startfile(files)





def centrar(ventana):
    """Centra cualquier ventana dependiendo el tamaño de la pantalla."""
    ventana.geometry('1066x600') # fija el tamaño de la ventana
    ventana.resizable(False, False) # indica que la ventana no se puede redimensionar
    ventana.update_idletasks() 
    w = ventana.winfo_width() # obtiene el ancho de la pantalla
    h = ventana.winfo_height() # obtiene el alto de la pantalla
    extraW = ventana.winfo_screenwidth()-w
    extraH = ventana.winfo_screenheight()-h
    ventana.geometry("%dx%d%+d%+d" % (w,h,extraW/2,extraH/2)) # posiciona la ventana





def next(ventana):
    """Minimiza la ventana pasada"""
    ventana.withdraw()


def back(cerrar, abrir):
    """Retrocede a la ventana anterior"""
    # se destruye la ventana anterior
    cerrar.destroy()
    # se abre la ventana anterior
    abrir.deiconify()


def openfiles(pressed):
    """Permite abrir la base de datos o directamente la carpeta de reportes."""
    if pressed == 'abrirbase':
        # abre el archivo donde estan todoso los estudiantes con Beneficio del comedor
        os.startfile(r'C:\SistemaComedor\Comedor.xlsx')
    else:
        abrir_reportes()


def Second_win(selection):
    """Crea un nuevo Toplevel dependiendo el boton que se presione"""
    global entry, CedulasRegistradas, top

    def ventana(img):
        """Muestra la ventana dependiendo del grupo que pertenezca la persona\n
        rojo -> Ya comió\n
        amarrillo -> Tiquete\n
        verde -> Puede Pasar"""
        
        global image2
        
        entry.delete(0, END) # Borra todo el texto que se introduce en el cuadro de texto
        color = Toplevel(top)

        image2 = PhotoImage(file=img); image2 = image2.subsample(1,1) # Inserta la imagen correspondiente
        label = Label(color, image=image2)   
        label.place(x=0,y=0,relwidth=1.0,relheight=1.0)
        
        centrar(color)

        def ocultar():
            """Destruye una ventana de colores."""
            color.destroy()
        color.after(2000,ocultar)






    def Enter(event):
        """Registra cada una de las cédulas que se van introduciendo"""
        CedulasRegistradas.clear() # Evita que los datos se dupliquen en el registro
        cargar_json()
        NumeroDeCedula = (entry.get()).upper() 
        try:
            if NumeroDeCedula.isalpha() == False and NumeroDeCedula.isalnum() == True:
                if NumeroDeCedula not in CedulasRegistradas:
                    if NumeroDeCedula in PersonasDelComedor:
                        # Si la persona está dentro del archivo.
                        UsaronElComedor[NumeroDeCedula] = PersonasDelComedor[NumeroDeCedula]
                        colorw = resource_path(r'Files\green.png')
                        ventana(colorw)
                    else:
                        # Si la persona no esta dentro del archivo.
                        colorw = resource_path(r'Files\gold.png')
                        ventana(colorw)

                        UsaronElComedor[NumeroDeCedula] = ('Estudiante Regular', 'NA', 'Ventas')
                    CedulasRegistradas.append(NumeroDeCedula)
                else:
                    # Si la cédula ya fue registrada.
                    colorw = resource_path(r'Files\red.png')
                    ventana(colorw)

                # Guarda el registro
                editar_json(); GuardarRegistro()
                UsaronElComedor.clear()# se resetea el diccionario

            elif NumeroDeCedula.isalpha() == True:
                messagebox.showerror('Error', 'Tiene que estar en formato alfanúmerico.')
                
            elif '' in NumeroDeCedula:      
                messagebox.showerror('Error', 'No se permiten espacios y otros caracteres que no sean letras o números.')

            entry.delete(0, END)

        except:
            messagebox.showerror("Error", "Un error inesperado acaba de ocurrir :(")
            entry.delete(0, END)






    if VERIFICACION_ARCHIVO('Comedor.xlsx') is True:
        global top
        next(root)

        if selection == 'InsertarCédulas':
            # ventana secundaria
            top = Toplevel(root)

            # Fondo de la ventana secundaria.
            img = resource_path(r'Files\INS.png')
            image = Image.open(img)
            tk_image = ImageTk.PhotoImage(image)
            Label(top, image = tk_image ).pack() # establece el fondo de la ventana

            # Cuadro de texto.
            entry = Entry(top, justify='center', font = ('Comic Sans Ms', 40),width= 21, borderwidth=0)
            entry.bind('<Return>', Enter) # cuando se presione la tecla enter el programa ejecutará la función Enter
            entry.place(x = 195, y = 266)

            
        elif selection == 'AbrirBaseDeDatos':
            # Ventana secundaria
            top = Toplevel(root)

            # Fondo de la ventana secundaria.
            img = resource_path(r'Files\archivos.png')
            imagen = PhotoImage(file = img)
            Label(top, image= imagen, bd = 0).pack() # establece el fondo de la ventana

            # Botones
            Reportes = Button(top)
            Bases = Button(top)

            # Configuración de los botones
            Reportes.config(
                            text= 'Revisar\nreportes', 
                            font = Font_tuple, 
                            width= 15, 
                            background='#ffffff', 
                            activebackground='#ffffff', 
                            foreground='black', 
                            activeforeground='black', 
                            borderwidth= 0, 
                            relief='raised', 
                            overrelief='sunken',
                            command=lambda: openfiles('revisar reportes')
            )

            Bases.config(
                        text= 'Abrir\nbase de datos', 
                        font = Font_tuple, 
                        width= 15, 
                        background='#ffffff', 
                        activebackground='#ffffff', 
                        foreground='black', 
                        activeforeground='black', 
                        borderwidth= 0, 
                        relief='raised', 
                        overrelief='sunken', 
                        command=lambda: openfiles('abrirbase')
            )

            # Poscicionamientos de los botones
            Reportes.place(x = 199, y = 270)
            Bases.place(x = 627, y = 270)

        # Botones
        cerrar = Button(top)
        atras = Button(top)

        # Configuración de los botones
        atras.config(

                    text= 'Atrás', 
                    font = 'Comic_Sans_Ms 9', 
                    width= 20, 
                    background='#f0f0f0', 
                    activebackground='#f0f0f0', 
                    foreground='black', 
                    activeforeground='black', 
                    borderwidth= 2,
                    relief='raised', 
                    overrelief='sunken', 
                    command=lambda: back(top, root)
        )

        cerrar.config(

                    text= 'Cerrar', 
                    font = 'Comic_Sans_Ms 9', 
                    width= 20, 
                    background='#f0f0f0', 
                    activebackground='#f0f0f0', 
                    foreground='black', 
                    activeforeground='black', 
                    borderwidth= 2,
                    relief='raised', 
                    overrelief='sunken', 
                    command=lambda: exit()
        )



        # Poscicionamiento de los botones
        atras.place(x = 20, y = 570 )
        cerrar.place(x = 900, y = 570 )

        
        centrar(top)
        top.protocol('WM_DELETE_WINDOW', lambda: back(top, root))
        top.mainloop()


# INTERFAZ GRÁFICA.
def main():

    global root, entry, Bases, BCedulas


    root = Tk()
    root.title('Comedor')

    #Icono
    icon = resource_path(r'Files\icono.ico') 
    root.wm_iconbitmap(True, icon) # establece el icono de la ventana

    # Imagen de fondo.
    img = resource_path(r'Files\main.png') 
    imagen = PhotoImage(file = img) 
    Label(root, image= imagen, bd = 0).pack() # establece el fondo de la ventana



    centrar(root)

    # Botones.

    Bases = Button(root)
    Insertar = Button(root)



    # configuración de los botones.
    Bases.config(
        
            text= 'Revisar\narchivos', 
            font = Font_tuple, 
            width= 15, 
            background='#ffffff',
            activebackground='#ffffff', 
            foreground='black', 
            activeforeground='black', 
            borderwidth= 0, 
            relief='raised', 
            overrelief='sunken', 
            command=lambda: Second_win('AbrirBaseDeDatos')
        
    )


    Insertar.config(
        
            text= 'Ingresar\ncédulas', 
            font = Font_tuple, 
            width= 15, 
            background='#ffffff', 
            activebackground='#ffffff', 
            foreground='black', 
            activeforeground='white', 
            borderwidth= 0, 
            relief='raised',
            overrelief='sunken', 
            command=lambda: Second_win('InsertarCédulas'),
            
    )






    # posicionamiento de los botonoes.
    Bases.place(x = 200, y = 270)
    Insertar.place(x = 627, y = 270)

    # TODO: Añadir el manual para que sirva este botón.
    # Manual = resource_path('Manual/Manual.pdf')
    # man = Button(root)
    # man.config(text= 'Manual', font = 'Comic_Sans_Ms 9', width= 20, background='black', activebackground='black', foreground='white', activeforeground='white', borderwidth= 2, relief='raised', overrelief='sunken', command=lambda: os.startfiles(Manual))
    # man.place(x = 20, y = 570 )

    root.mainloop()

if __name__ == "__main__":
    os.makedirs(r'C:\SistemaComedor', exist_ok=True) # Crea la carpeta donde se guardarán los reportes
    main()