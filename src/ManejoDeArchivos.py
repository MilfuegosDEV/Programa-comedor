import json
import os
from datetime import datetime
from typing import Literal

import pandas as pd
from openpyxl import Workbook, load_workbook


class __generarCarpetas:
    '''Se encarga de crear todos las carpetas necesarias para el funcionaiento del programa.'''

    def __init__(self, directory: str = "\\SistemaComedor") -> None:
        '''Crea las carpetas donde se guardarán la base de datos y los reportes generados por el programa.'''
        self._directory: str = directory
        '''Carpeta principal.'''
        os.makedirs(self._directory, exist_ok=True)  # crea la carpeta.

    def mksubdirectory(self) -> None:
        '''Genera las carpetas donde se guardarán los reportes.'''
        self.__months: list = ["Enero",
                               "Febrero",
                               "Marzo",
                               "Abril",
                               "Mayo",
                               "Junio",
                               "Julio",
                               "Agosto",
                               "Setiembre",
                               "Octubre",
                               "Noviembre",
                               "Diciembre"]  # Lista de meses.
        self._thisMonth = self.__months[int(
            datetime.today().strftime('%m'))-1]  # Mes actual

        self._reportes: str = self._directory + "\\Reportes"
        '''Carpeta de reportes.'''
        self._carpetaMensual: str = self._reportes + \
            f"\\{datetime.today().strftime('%Y')}" + f"\\{self._thisMonth}"
        '''Carpeta anual con un subdirectorio que indica el mes.'''
        self._cache: str = self._reportes + "\\Cache"
        '''Carpeta que guarda el cache generado diariamente.'''

        os.makedirs(self._carpetaMensual, exist_ok=True)

        try:
            os.system(f'attrib +h {self._cache}')  # Windows
        except:
            os.rename(self._cache, self._reportes + "\\.Cache")  # Linux


class leerBaseDeDatos(__generarCarpetas):
    '''Verifica la información contenida en la base de datos.'''

    def __init__(self, directory: str = "\SistemaComedor",
                 file: str = "Comedor.xlsx") -> None:
        super().__init__(directory)

        self.__file: str = file
        '''Excel donde se extraerá la información.'''

    @property
    def info(self) -> dict | str | None:
        '''Recolecta la información contenida en la base de datos.'''
        __info: dict = {}
        try:
            # Abre la base de datos
            df = pd.read_excel(self._directory+"\\"+self.__file)
            for row in df.itertuples(index=True):
                # si hay menos datos en la fila de los cuales son requeridos
                # el programa indicará la fila en la cual hay un error.
                if 'nan' in str(row[0:3]) or ((str(row[1]).strip()).isalpha() == True or (str(row[1]).strip()).isalnum() == False) or str(row[2]).isnumeric() == True:
                    # Requisitos:
                    # No pueden haber celdas vacías en la cedula o nombre.
                    # La cédula no pueden ser solamente letras, debe ser una combinación entre letras y números o solamente o números.
                    # El número de cédula no puede contener espacios o algún carácter especial.
                    # El nombre no puede contener ningun número.
                    # No pueden haber filas vacias entre filas.
                    self.__fila: int = row[0] + 2
                    return f"Se ha encontrado un error en la fila {self.__fila}, por favor reviselo, esta debe contener:\nCédula, Nombre completo y sección (opcional)."
                else:
                    try:
                        __info[str(row[1].upper()).strip()] = {
                            "Nombre": row[2].strip(), "Seccion": row[3]}
                    except AttributeError:
                        __info[str(row[1]).strip()] = {
                            "Nombre": row[2].strip(), "Seccion": row[3]}
            if df.empty == True:  # Si el archivo excel esta vacio.
                return None
            return __info
        except FileNotFoundError:
            return f"No se ha encontrado el archivo {self.__file}"


class generadorDeReportes(__generarCarpetas):
    '''Crea los reportes de ingresos'''

    def __init__(self, data: dict,
                 reportFile: str = f"ReporteDelMes.xlsx",
                 directory: str = "\SistemaComedor") -> None:
        super().__init__(directory)
        self.mksubdirectory()

        self.__reportFile: str = reportFile
        '''Nombre del archivo reporte'''
        self.__data: dict = data
        '''información recolectada del estudiante que fue ingresado.'''
        self.__hoy = pd.to_datetime(datetime.today().strftime(
            '%d-%m-%y %H:%M:%S'), dayfirst=True)

    def GuardarRegistro(self) -> Literal[True] | str:
        '''Edita el reporte.'''
        try:
            # Carga el reporte.
            xlF = load_workbook(self._carpetaMensual +
                                f"\\{self.__reportFile}")
            ws = xlF.active  # Hoja activa
            i = ws.max_row
            i += 1  # encuentra la última linea del archivo

            # Se edita el reporte.
            ws.protection.disable()
            for id, infoIngreso in self.__data.items():
                # {"123456": {"Nombre": Pepito, "Seccion": "7-1"}}
                # id = "123456"
                # infoIngresos = {"Nombre": Pepito, "Seccion": "7-1"}
                try:
                    ws[f'A{i}'] = int(id)  # Número de cédula
                except ValueError:
                    ws[f'A{i}'] = id  # Número de cédula
                finally:
                    ws[f'B{i}'] = infoIngreso["Nombre"]  # Nombre completo
                    ws[f'C{i}'] = infoIngreso['Seccion']  # Sección
                    ws[f'D{i}'] = self.__hoy  # Fecha con hora
            ws.protection.enable()

            # Guarda el reporte editado
            xlF.save(self._carpetaMensual + f'\\{self.__reportFile}')
            xlF.close()  # Cierra el reporte
            return True
        except FileNotFoundError:
            '''Genera reporte.'''
            wb = Workbook()
            '''Libro de reporte.'''
            ws = wb.active  # Hoja activa.
            # Contraseña de los reportes.
            ws.protection.password = 'MySecretKey'
            ws.title = 'Reporte'

            # Titulo de cada columna.
            ws.append({1: 'Cédula',
                       2: 'Nombre completo',
                       3: 'Sección',
                       4: 'Fecha'})

            # Ancho de cada columna
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20

            # Se guarda el archivo vacío
            wb.save(self._carpetaMensual + f"\\{self.__reportFile}")
            wb.close()  # Cierra el archivo.

            # edita el archivo con la información ingresada.
            return self.GuardarRegistro()
        except PermissionError:
            return f'Por favor cierre el archivo {self.__reportFile} antes de continuar'


class generadorDeCache(__generarCarpetas):
    '''Genera datos para verificar si la persona ya utilizó el comedor.'''

    def __init__(self, directory: str = "\SistemaComedor") -> None:
        super().__init__(directory)
        self.mksubdirectory()

    @property
    def cache(self) -> dict:
        '''Carga la información'''
        try:
            with open(self._cache+"\\"+self._thisMonth+".json", "r") as f:
                return json.load(f)
        except (FileNotFoundError):
            self.editar({})
            return self.cache

    def editar(self, infoIngreso: dict) -> bool:
        '''Edita la información.'''
        try:
            with open(self._cache+"\\"+self._thisMonth+".json", "w", encoding="utf-8") as f:
                json.dump(infoIngreso, f, indent=4)
                f.close()
            return True
        except FileNotFoundError:
            return False
