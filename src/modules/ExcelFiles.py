from datetime import datetime
import os, pandas as pd, shutil, platform
from openpyxl import Workbook, load_workbook
from tkinter import messagebox, filedialog as fd

class xlFiles:
    cache = '' # carpeta de los registros diarios.
    info = {} # Personas dentro del archivo

    def __init__(self, dir: str) -> None:
        # carpetas donde se guardarán los archivos
        self.__dir = dir
        self.__foldername = self.__dir +"\\Reports"
        if platform.system() == 'Windows':
            self.cache = self.__foldername + '\\Cache'
            os.makedirs(self.cache, exist_ok= True)
            os.system(f'attrib +h {self.cache}')
        elif platform.system() == 'Linux':
            self.cache = self.__foldername + '\\.Cache'
            os.makedirs(self.cache, exist_ok= True)

    def VerificacionDeDatos(self, Archivo: str) -> bool:
        # Verifica si los datos del archivo corresponden o no.
        self.__filename = self.__dir + f"\\{Archivo}"
        try:
            df = pd.read_excel(self.__filename)
            # Tomando los datos del archivo
            for row in df.itertuples(name = None, index = False):
                try:
                    self.info[str(row[0].upper())] = row[1::]
                except:
                    self.info[str(row[0])] = row[1::]
            for k, v in self.info.items():
                if k.isalpha() == True or k.isalnum() == False or v[0].isnumeric() == True or df.isnull().values.any() == True or len(v) > 3:
                    self.info.clear()
                    break
            if self.__IsNotEmpty(self.info) == True:
                return True
        except FileNotFoundError as e:
            messagebox.showerror('FileNotFoundError', f'{e}\nPor favor presione abrir para mover el archivo.')
            source = fd.askopenfilename(title = 'Mover',filetypes=(('Excel files', '.xlsx'), ('All', '.')))
            destination = self.__dir
            try:
                shutil.move(source, destination)
            except shutil.Error:
                pass
            return False
        except ValueError:
            os.startfiles(self.__filename)
            messagebox.showwarning(f'Formato incorrecto', 'El archivo debe tener el siguiente formato\nCédula, Nombre completo, Sección, Grupo')
            return False
        
        except Exception as e:
            messagebox.showerror('Error Inesperado', f'{e}')

    def __IsNotEmpty(self, data_structure: dict) -> bool:
        if (len(data_structure) > 0):
            return True
        else:
            os.startfile(self.__filename)
            messagebox.showwarning('Formato incorrecto', 'El archivo debe tener el siguiente formato\nCédula, Nombre completo, Sección, Grupo')
            return False

    def GuardarRegistro(self, info: dict):
        __hoy = pd.to_datetime(datetime.today().strftime('%d-%m-%y %H:%M:%S'), dayfirst=True)
        __actual = datetime.today().strftime('%m-%y')
        while True:
            try:
                archivo = load_workbook(self.__foldername +f"\\{__actual}.xlsx")
                ws = archivo.active
                i = ws.max_row; i += 1; # encuentra la última linea del archivo
                for k, v in info.items():
                    """
                    Cédula         | Nombre completo              | Sección | Grupo                 | Fecha
                    305550820        Juan Daniel Luna Cienfuegos      11-1    Estudiante Regular      1-7-2022    
                    """
                    try:
                        ws[f'A{i}'] = int(k) # Número de cédula
                    except ValueError:
                        ws[f'A{i}'] = k # Número de cédula

                    ws[f'B{i}'] = v[0] # Nombre completo 
                    ws[f'C{i}'] = v[1] # Sección
                    ws[f'D{i}'] = v[2] # Grupo
                    ws[f'E{i}'] = __hoy # Fecha con hora
                    i += 1 # avanza a la siguiente linea.
                archivo.save(self.__foldername + f'\\{__actual}.xlsx')
                break
            except FileNotFoundError:
                wb = Workbook()
                ws = wb.active
                ws.title = 'Registro'
                ws.append({1:'Cédula',
                           2:'Nombre completo', 
                           3: 'Sección',
                           4: 'Grupo',
                           5: 'Fecha'})
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 50
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 20
                wb.save(self.__foldername +f"\\{__actual}.xlsx")
                continue
            except PermissionError as e:
                messagebox.askretrycancel(f'PermissionError', f'{e}\nPor favor cierre el archivo antes de continuar')
                continue
            except ValueError:
                messagebox.showwarning(f'Formato incorrecto', 'El archivo debe tener el siguiente formato\nCédula, Nombre completo, Sección, Grupo')
                os.startfile(self.__filename)
                continue      

if __name__ == "__main__": 
    __xlF = xlFiles("C:\\Users\\juand\\OneDrive\\Escritorio\\Pruebas") # Se especifica la carpeta en la cual se quiere trabajar
    __xlF.VerificacionDeDatos('Comedor.xlsx') # se especifica el nombre del archivo con el cual se quiere trabajar
    __xlF.GuardarRegistro({305550820: ('Juan Daniel Luna Cienfuegos', '11-1', 'Regular')}) # Prueba de guardado de registro.
    print(__xlF.info) # datos recolectados
