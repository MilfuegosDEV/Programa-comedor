from datetime import datetime
import os, pandas as pd, shutil, json, time
from openpyxl import Workbook, load_workbook
from tkinter import messagebox, filedialog as fd

class xlFiles:

    def __init__(self, dir: str, Archivo: str) -> None:
        """
        Verificación del archivo xlsx del cuál se alojan los datos del estudiante y generación
        del reporte mensual de los estudiantes que asistieron al comedor
        """
        
        self.__Archivo = Archivo
        """Nombre del archivo"""
        self.cache = '' 
        """Carpeta con las cédulas registradas durante el día."""
        self.info = {}  

        self.filename = ''
        """Base de datos"""
        self.foldername = ''
        """Carpeta de reportes"""
        
        self.actual = "Test"

        self.__crearCarpetas(dir, Archivo)

    def __crearCarpetas (self, dir: str, Archivo: str):
        """Creación de la carpeta donde se alojarán los archivos de los estudiantes
    
        AMBAS CARPETAS EN EL CASO DE EXISTIR NO SE MODIFICAN.

        Args:
            dir (str): Carpeta con los archivos requeridos por el programa y proporcionados por este mismo.
            Archivo (str): Carpeta de los registros diarios.
        """
        
        self.__dir = dir
        self.filename = self.__dir + f"\\{Archivo}"
        self.foldername = self.__dir + f"\\Reportes\\{datetime.today().strftime('%Y')}"

        # Se se indica el nombre del mes.
        if datetime.today().strftime("%m") == '01':
            self.foldername = self.foldername + "\\Enero"
        elif datetime.today().strftime("%m") == '02':
            self.foldername = self.foldername + "\\Febrero"
        elif datetime.today().strftime("%m") == '03':
            self.foldername = self.foldername + "\\Marzo"
        elif datetime.today().strftime("%m") == '04':
            self.foldername = self.foldername + "\\Abril"
        elif datetime.today().strftime("%m") == '05':
            self.foldername = self.foldername + "\\Mayo"
        elif datetime.today().strftime("%m") == '06':
            self.foldername = self.foldername + "\\Junio"
        elif datetime.today().strftime("%m") == '07':
            self.foldername = self.foldername + "\\Julio"
        elif datetime.today().strftime("%m") == '08':
            self.foldername = self.foldername + "\\Agosto"
        elif datetime.today().strftime("%m") == '09':
            self.foldername = self.foldername + "\\Setiembre"
        elif datetime.today().strftime("%m") == '10':
            self.foldername = self.foldername + "\\Octubre"
        elif datetime.today().strftime("%m") == '11':
            self.foldername = self.foldername + "\\Noviembre"
        elif datetime.today().strftime("%m") == '12':
            self.foldername = self.foldername + "\\Diciembre"

        try:
            self.cache = self.foldername + '\\Cache'
            self.foldername 
            # se crea la carpeta.
            os.makedirs(self.cache, exist_ok= True)
            os.system(f'attrib +h {self.cache}')
        except:
            exit()


    def VerificacionDeDatos(self) -> bool:
        """Verifica los datos del archivo los cuales deben estan en el siguente formato:
        
            | Cédula | Nombre completo | Sección

        En el caso de que de que el archivo no este en la carpeta indicada este abrirá el administrador de archivos 
        para que podamos movamos el archivo a la carpeta solicitada.

        En el caso de que haya una fila con menos datos de los solicitados el programa o que en esta fila
        hayan datos que no esten en el formato correcto. El programa mostrará una ventana de dialogo donde se mostrará la fila
        en la cual se encuentra el error, de igual forma si el archivo esta vacio mostrara otra ventana de dialogo indicando el error.

        Returns:
            bool: Retornará True en el caso de que toda la información del archivo sea correcta, de lo contrario false.
        """

        try:
            df = pd.read_excel(self.filename)
            # Tomando los datos del archivo
            
            for row in df.itertuples(index = True):
                # si hay menos datos en la fila de los cuales son requeridos el programa indicará la fila en la cual hay un error.
                
                if 'nan' in str(row[0:3]) or ((str(row[1]).strip()).isalpha() == True or (str(row[1]).strip()).isalnum() == False) or str(row[2]).isnumeric() == True:
                    # Requisitos:

                    # No pueden haber celdas vacías en la cedula o nombre.
                    # La cédula no pueden ser solamente letras, debe ser una combinación entre letras y números o solamente o números.
                    # El número de cédula no puede contener espacios o algún carácter especial.
                    # El nombre no puede contener ningun número.

                    # Bug: Cuando hay celdas extras con datos extras en otras filas el programa las ignora.
                    self.__fila = row[0] + 2
                    messagebox.showerror('Error en fila', f'Revise la fila {self.__fila}\nEl archivo debe tener el siguiente formato\nCédula, Nombre completo y sección.')
                    os.startfile(self.filename)
                    return False
                
                else:
                    try:
                        self.info[str(row[1].upper()).strip()] = row[2:4]
                    except AttributeError:
                        self.info[str(row[1]).strip()] = row[2:4] 
            
            if df.empty == True:
                # En el caso de que el archivo este vacio
                messagebox.showerror('Archivo vacio', f'El archivo no puede estar vacio.\nEl archivo debe tener el siguiente formato\nCédula, Nombre completo y sección.')
                os.startfile(self.filename)
                return False
            
            # En el caso de que se realice todo con normalidad.
            return True

        except FileNotFoundError:
            # En el caso de que no se encuentre el archivo
            messagebox.showerror('FileNotFoundError', f'No se ha encontrado el archivo {self.__Archivo}\nPor favor presione abrir para copiar el archivo.')
            source = fd.askopenfilename(title = 'Mover',filetypes=(('Excel files', '.xlsx'), ('All', '.')))
            
            __nombre = (source).split("/")
            __nombre = (__nombre[-1])
            __nombre = __nombre.split('.')
            __nombre = __nombre[0]

            __archivo = self.__Archivo.split('.')
        

            destination = self.__dir

            if __nombre.title() == __archivo[0].title():
                try:
                    shutil.copy(source, destination)
                except shutil.Error:
                    pass
                return False
            else:
                messagebox.showwarning("Nombre incorrecto", f"El archivo debe tener el nombre de: {self.__Archivo}")
                return False
                

        except:
            messagebox.showerror('Error Inesperado', 'Ha ocurrido un error inesperado.')
            return False


    def GuardarRegistro(self, data: dict):
        """Guarda el registro de los ingresos de los becados al comedor.

        Args:
            data (dict): {Número cédula: [Nombre completo, sección]}
        """


        __hoy = pd.to_datetime(datetime.today().strftime('%d-%m-%y %H:%M:%S'), dayfirst=True)

        while True:
            try:
                archivo = load_workbook(self.foldername +f"\\{self.actual}.xlsx")
                ws = archivo.active
                i = ws.max_row; i += 1; # encuentra la última linea del archivo
                for k, v in data.items():
                    """
                    Cédula         | Nombre completo       | Sección
                    1111111111       Roberto Robles Gomez       8-2
                    """
                    __cedula = k
                    __nombreCompleto = v[0]
                    __seccion = v[1]

                    try:
                        ws[f'A{i}'] = int(__cedula) # Número de cédula
                    except ValueError:
                        ws[f'A{i}'] = __cedula # Número de cédula
                    finally:
                        ws[f'B{i}'] = __nombreCompleto # Nombre completo
                        ws[f'C{i}'] = __seccion # Sección 
                        
                        ws[f'D{i}'] = __hoy # Fecha con hora
                        i += 1 # avanza a la siguiente linea.
                
                archivo.save(self.foldername + f'\\{self.actual}.xlsx')
                break

            except FileNotFoundError:

                wb = Workbook()
                ws = wb.active
                ws.title = 'Registro'
                ws.append({1:'Cédula',
                           2:'Nombre completo',
                           3: 'Sección',
                           4: 'Fecha'})
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 50
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 20

                wb.save(self.foldername +f"\\{self.actual}.xlsx")
                continue

            except PermissionError as e:
                messagebox.askretrycancel(f'PermissionError', f'{e}\nPor favor cierre el archivo antes de continuar')
                continue

            except ValueError:
                messagebox.showwarning(f'Formato incorrecto', 'El archivo debe tener el siguiente formato\nCédula, Nombre completo y sección.')
                os.startfile(self.filename)
                continue      


class Temp:
    """
    Registro de los números de cédulas que han usado el comedor durante el dia
    - tempinfo: una lista con los números de cédula de las personas que utilizaron el comedor
    durante el día.
    """
    tempinfo = [] # info dentro del archivo json.
    __hoy = time.strftime('%d-%m-%y')
    def Cargar_info(self, CacheFolder):
        """
        Extrae la información que hay en el archivo de tipo .json
        en y la guarda en una lista.
        """
        try:
            with open(f'{CacheFolder}\\{self.__hoy}.json', 'r') as File:
                self.tempinfo = json.load(File)
        except FileNotFoundError:
            self.Editar_info(CacheFolder, self.tempinfo)

    def Editar_info(self, CacheFolder: str, data: list):
        """
        Edita el archivo de tipo .json con los números de cédula recolectados.

        - data: lista de números de cedulas.
        """
        try:
            with open(f'{CacheFolder}\\{self.__hoy}.json', 'w') as File:
                json.dump(data, File, indent=4)
                File.close()
        except FileNotFoundError:
            ask = messagebox.askokcancel('FileNotFoundError', 'Se supone que esto no debia suceder. Reinicie el programa, pero la información que fue recolectada durante el día ha sido perdida.')
            if ask == True:
                exit()
            else:
                pass


if __name__ == "__main__": 
    # Test
    __xlF = xlFiles("SistemaComedor", "Comedor.xlsx") # Se especifica la carpeta en la cual se quiere trabajar
    __xlF.VerificacionDeDatos() # se especifica el nombre del archivo con el cual se quiere trabajar
    __xlF.GuardarRegistro({11111111: ('Roberto Robles Gomez', '11-1')}) # Prueba de guardado de registro.
    print(__xlF.info) # datos recolectados