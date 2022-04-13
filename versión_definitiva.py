'''
La idea del programa es verificar mediante un código de barras las cédulas estan en el
alguna base de datos.

Para efectos de simplicidad trabajaremos con archivos excel ('.xlsx').
'''
# Módulos ('librerias')
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
import os
import openpyxl
import pandas as pd
import datetime
from shutil import move



# Variables.
PlanNacional = {}
PlanUsado = {}
becas = {}
BecasUsadas = {}
ventas = []
CédulasRegistradas = []
UltimaFila = 0 


# Programa
def mkdir(): # crea el directorio en la carpeta donde se guardarán los excel
    os.makedirs(r'C:\SistemaComedor', exist_ok=True)


    f = open(r"C:\SistemaComedor\ImportanteLeer.txt", "w", encoding= "utf-8")
    f.write("""POR FAVOR AÑADA LOS ARCHIVOS REQUERIDOS EN ESTA CARPETA

LOS CUALES DEBEN TENER TENER LOS SIGUIENTES NOMBRES:

Becados.xlsx (para las personas las cuales tienen beca del comedor)
PlanNacional.xlsx (para las personas que pertencen a plan nacional)

Si alguno de los dos archivos no tienen el nombre como esta indicado el programa continuará tirando error, además los ambos archivos
deberán tener el siguiente formato:
		
		A		  B
	1|  CEDULAS  |	       NOMBRES       	   |
	2| 305550820 | Juan Daniel Luna Cienfuegos |

Por Favor busque los archivos y guardelos en la carpeta C:\SistemaComedor 

En el caso de algun problema se pueden comunicar conmigo:
correo: juandaniel.lunacienfuegos@gmail.com
cuenta de Github: https://github.com/MilfuegosxD""")

    f.close()

def isnotEmpty(data_structure, file): # Revisa si los archivos están en el formato correcto o si no están vacios
    if len (data_structure) > 0:
        return True
    else:
        messagebox.showinfo('Formato incorrecto',f'''El archivo {file} debe tener el siguiente formato: 
                A                  B
        1| CEDULAS |     NOMBRES      |
        2| 30328349| Federico Lopez...|''')
        return False

# Gestión de archivos
def Verificación_de_archivos_de_Plan_Nacional():
    Archivo = 'PlanNacional.xlsx'
    try:
        wb = openpyxl.load_workbook(r'C:\SistemaComedor\PlanNacional.xlsx')
        ws = wb.active
        rows = ws.iter_rows(min_row = 2,
                            max_row= 100000, 
                            min_col = 1, 
                            max_col = 2
                            )
        for a, b in rows:
            PlanNacional[a.value] = b.value
        PlanNacional.pop(None)
        
        for clave in PlanNacional.keys():
            if str(clave).isnumeric() == False:
                PlanNacional.clear()
                break
            else:
                continue
            
        if isnotEmpty(PlanNacional, Archivo) == True:
            return True       
    except FileNotFoundError:
        messagebox.showerror('FileNotFoundError',f'Agregue el archivo {Archivo} al directorio C:\SistemaComedor')

        return False

def Verificación_de_archivos_de_Becados():
    Archivo = 'Becados.xlsx'
    try:
        wb = openpyxl.load_workbook(r'C:\SistemaComedor\Becados.xlsx')
        ws = wb.active
        rows = ws.iter_rows(min_row=2, 
                            max_row= 100000, 
                            min_col=1, 
                            max_col= 2
                            )

        for a, b in rows:
            becas[a.value] = b.value
        becas.pop(None)
        
        for clave in becas.keys():
            if str(clave).isnumeric() == False:
                becas.clear()
                break
            else:
                continue
        if isnotEmpty(becas, Archivo) == True:
            return True

    except FileNotFoundError:
        messagebox.showerror('FileNotFoundError',f'Agregue el archivo {Archivo} al directorio C:\SistemaComedor')
        return False

def GuardarRegistro():
    os.makedirs(r'C:\SistemaComedor\reportes', exist_ok=True)
    while True:
        try:
            fecha = str(datetime.datetime.today().strftime('%m-%y'))
            hoy = datetime.datetime.today().strftime('%d/%m/%y')
            hoy = pd.to_datetime(hoy, dayfirst= True,)
            archivo = openpyxl.load_workbook(r'C:\SistemaComedor\reportes\Reporte '+ fecha+'.xlsx')
            ws = archivo.active
            i = ws.max_row 
            i += 1
            ws.protection.disable() # deshabilita la protección del archivo para editarlo
            for cedula, nombre in PlanUsado.items():
                ws[f'A{i}']= cedula
                ws[f'B{i}']= nombre
                ws[f'C{i}']= 'Plan Nacional'
                ws[f'D{i}']= hoy
                i += 1
            for cedulab, nombreb in BecasUsadas.items():
                ws[f'A{i}'] = cedulab
                ws[f'B{i}'] = nombreb
                ws[f'C{i}'] = 'Becados'
                ws[f'D{i}'] = hoy
                i+= 1

            for cedulasv in ventas:
                ws[f'A{i}'] = cedulasv
                ws[f'B{i}'] = 'Desconocido'
                ws[f'C{i}'] = 'Ventas'
                ws[f'D{i}'] = hoy
                i += 1
            ws.protection.enable() # rehabilita la protección del archivo para que no pueda ser modificado. 
            archivo.save(r'C:\SistemaComedor\reportes\Reporte '+ fecha+'.xlsx')
            break

        except FileNotFoundError:
            fecha = str(datetime.datetime.today().strftime('%m-%y'))
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Registro'
            ws.append({1:'CEDULA', 2:'NOMBRE', 3:'GRUPO', 4:'FECHA'})
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            wb.save(fr'C:\SistemaComedor\reportes\Reporte '+fecha+'.xlsx')
            continue
        except PermissionError:
            messagebox.showwarning('Reporte abierto',f"Por Favor cierre el archivo antes de continuar 'Reporte {fecha}.xlsx'")
            continue

# Intérfaz Gráfica


# Funciones de la interfaz
def on_closing():
    if messagebox.askyesno('Salir', '¿Salir del programa?'):
        root.destroy()
mkdir()

root = Tk()
root.title('Programa comedor')
root.geometry('500x300')
root.config(bg = '#380106')
root.resizable(0,0)

columna1 = Label(root)
columna1.config(bg = 'black', 
                height= 18
                )
columna1.place(x = 0, y = 12)


columna2 = Label(root)
columna2.config(bg = 'black', 
                height= 18
                )
columna2.place(x = 494, y = 12)

titulo = Label(root)
titulo.config(bg = 'white',
            fg = 'black',
            font = 'Times_new_roman 20',
            text = 'Menú',
            width  = 20,
            )
titulo.place(x = 90, y = 40)

# botón para el mantenimiento
boton1 = Button(root)
boton1.config(bg = 'black',
            activebackground= 'black',
            fg = 'white',
            activeforeground = 'white',
            font = 'Times_new_roman 15',
            text = 'Revisar\narchivos',
            width = 15,
            borderwidth= 4,
            command=lambda: Mantenimiento()
            )
boton1.place(x = 50, y = 140)
# Botón para registrar cédulas
boton2 = Button(root)
boton2.config(bg = 'black',
            activebackground= 'black',
            fg = 'white',
            activeforeground = 'white',
            font = 'Times_new_roman 15',
            text = 'Registrar\ncédulas',
            width = 15,
            borderwidth= 4,   
            command=lambda: Registrar_Cedulas()
            )
boton2.place(x = 270, y = 140)

root.protocol("WM_DELETE_WINDOW", on_closing)
'''
La siguiente parte del programa consistirá 
en todas la funciones que tienenlas ventanas hijas.
'''

# Funciones del botón 2
def Registrar_Cedulas():
    global Verificación_de_archivos_de_Becados, Verificación_de_archivos_de_Plan_Nacional
    
    def insertar_cedulas():
            """En esta parte del programa se le pide a la personas que inserte su número de cédulas
            Luego lo que hace es validar si el número de cudlas estan en:
            1. Plan Nacional
            2. Personas Becadas
            3. Personas las cuales deben comprar el almuerzo""" 

            # Primer menú
            '''En esta parte se ingresan los números de cédulas y el programa los busca en distintos diccionarios.'''

            '''El programa tirará error si:
                1. El número de cédula esta vacío 
            2. Si el número de cédula tiene letras'''
            try:
                NumeroDeCedula = int(entry.get())

                # En esta parte se verifica si la persona ha comido 
                '''Esta parte es necesaria porque hay veces que las personas ya han comido y quieren prestar su carnet a otras personas
                para que la otra personas coma y por lo tanto no este programa no tendría sentido.'''

                if NumeroDeCedula in CédulasRegistradas:
                    colores('brown2','Ya comió')
                    entry.delete(0,END)

                # En el caso de que la personas no haya comido

                elif NumeroDeCedula == 0:
                    on_closing()
                    
                else:
                    if int(NumeroDeCedula) in PlanNacional.keys():
                        colores('light green', 'Es de Plan nacional')
                        PlanUsado[NumeroDeCedula] = PlanNacional[NumeroDeCedula]
                        CédulasRegistradas.append(NumeroDeCedula)
                        


                    elif int(NumeroDeCedula) in becas.keys():
                        colores('light green', 'Es becado')
                        BecasUsadas[NumeroDeCedula]= becas[NumeroDeCedula]
                        CédulasRegistradas.append(NumeroDeCedula)



                    elif int(NumeroDeCedula) not in becas.keys() and NumeroDeCedula not in PlanNacional.keys() and NumeroDeCedula != 0:
                        colores('gold', 'Tiene que pagar')
                        ventas.append(NumeroDeCedula)
                        CédulasRegistradas.append(NumeroDeCedula)
                    entry.delete(0,END)

            except ValueError:
                messagebox.showinfo('Valores no admitidos', 'Las cédulas no llevan letras.')
            except NameError:
                messagebox.showinfo('Valores no admitidos', 'Las cédulas no llevan letras.')
            except KeyError:
                messagebox.showinfo('Valores no admitidos', 'Las cédulas no llevan letras.')
            except SyntaxError:
                messagebox.showinfo('Valores no admitidos', 'Las cédulas no llevan letras.')

    def colores (color, text):
        colores = Toplevel(child)
        colores.geometry('500x300')
        colores.config(bg = color)
        colores.resizable(0,0)
        columna1 = Label(colores)
        columna1.config(bg = 'black', 
                        height= 18
                        )
        columna1.place(x = 0, y = 12)


        columna2 = Label(colores)
        columna2.config(bg = 'black', 
                        height= 18
                        )
        columna2.place(x = 494, y = 12)
        label = Label(colores)
        label.config(bg = 'black',
                fg = 'white',
                font = 'Times_new_roman 20',
                text = text,
                width  = 22,
                height= 1
                )
        label.place(x = 68, y = 95)

        def ocultar():
            colores.destroy()
        colores.after(1000,ocultar)

    def guardar(event):
        insertar_cedulas()

    def on_closing():

        GuardarRegistro()
        BecasUsadas.clear()
        PlanUsado.clear()
        ventas.clear()
        child.destroy()
        root.deiconify()
    

    a = Verificación_de_archivos_de_Plan_Nacional()
    b = Verificación_de_archivos_de_Becados()

    if a is True and b is True:
        root.withdraw()
        child = Toplevel(root)
        child.geometry('500x300')
        child.config(bg = '#380106')
        child.resizable(0,0)
        columna1 = Label(child)
        columna1.config(bg = 'black', 
                        height= 18
                        )
        columna1.place(x = 0, y = 12)


        columna2 = Label(child)
        columna2.config(bg = 'black', 
                        height= 18
                        )
        columna2.place(x = 494, y = 12)
        label = Label(child)
        label.config(bg = 'black',
                fg = 'black',
                font = 'Times_new_roman 20',
                text = '    ',
                width  = 22,
                height= 1
                )
        label.place(x = 68, y = 95)
        entry = Entry(child)
        entry.config(bg = 'white',
                fg = 'black',
                font = 'Times_new_roman 25',
                width  = 20,
                justify='center'
                )
        entry.place(x = 75, y = 101)
        entry.bind("<Return>", guardar)
        boton = Button(child)
        boton.config(
            background='black',
            activebackground='black',
            foreground='white',
            activeforeground='white',
            borderwidth=4,
            text = 'Insertar',
            font = 'Times_new_roman 15',
            width= 15,
            command=lambda:insertar_cedulas()
        )
        boton.place(x= 155, y = 170)
    else:
        os.startfile('C:\SistemaComedor\ImportanteLeer.txt')
        source = filedialog.askopenfilename(title = 'Mover',filetypes=(('Excel files', '.xlsx'), ('Excel files', '.xlsx')))
        destination = 'C:\SistemaComedor'
        move(source, destination)



    child.protocol('WM_DELETE_WINDOW', on_closing)
# Funciones del botón 1
def Mantenimiento():
    def abrir_archivo():
        archivo = filedialog.askopenfilename(title= 'Abrir reporte',initialdir=r'C:\SistemaComedor\reportes', filetypes=(('Excel files','*.xlsx'),('All files', '*.*')))
        os.startfile(archivo)

    def bases_de_datos():
        archivo = filedialog.askopenfilename(title= 'Abrir base de datos', initialdir = r'C:\SistemaComedor', multiple = True, filetypes=(('Excel files','*.xlsx'),('All files', '*.*')))
        for files in archivo:
            os.startfile(files)
    def on_closing():
        child.destroy()
        root.deiconify()


    a = Verificación_de_archivos_de_Plan_Nacional()
    b = Verificación_de_archivos_de_Becados()
    
    if a is False or b is False:
        os.startfile('C:\SistemaComedor\ImportanteLeer.txt')
        source = filedialog.askopenfilename(title = 'Mover',filetypes=(('Excel files', '.xlsx'), ('Excel files', '.xlsx')))
        destination = 'C:\SistemaComedor'
        move(source, destination)

    elif a is True and b is True:
        root.withdraw()
        child = Tk()
        child.title('Programa comedor')
        child.geometry('500x300')
        child.config(bg = '#380106')
        child.resizable(0,0)

        columna1 = Label(child)
        columna1.config(bg = 'black', 
                        height= 18
                        )
        columna1.place(x = 0, y = 12)


        columna2 = Label(child)
        columna2.config(bg = 'black', 
                        height= 18
                        )
        columna2.place(x = 494, y = 12)

        titulo = Label(child)
        titulo.config(bg = 'white',
                    fg = 'black',
                    font = 'Times_new_roman 20',
                    text = 'Archivos',
                    width  = 20,
                    )
        titulo.place(x = 90, y = 40)
        # Botón para revisar la base de datos
        boton1 = Button(child)
        boton1.config(bg = 'black',
                    activebackground= 'black',
                    fg = 'white',
                    activeforeground = 'white',
                    font = 'Times_new_roman 15',
                    text = 'Revisar\nbase de datos',
                    width = 15,
                    borderwidth= 4,
                    command=lambda: bases_de_datos()
                    )
        boton1.place(x = 50, y = 140)
        # Botón para revisar los reportes
        boton2 = Button(child)
        boton2.config(bg = 'black',
                    activebackground= 'black',
                    fg = 'white',
                    activeforeground = 'white',
                    font = 'Times_new_roman 15',
                    text = 'Revisar\nReportes',
                    width = 15,
                    borderwidth= 4,
                    command=lambda: abrir_archivo()
                    )
        boton2.place(x = 270, y = 140)
        child.protocol('WM_DELETE_WINDOW', on_closing)














root.mainloop()





