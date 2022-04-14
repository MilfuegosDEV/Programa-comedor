"""
La idea es verificar mediante un código de baras las cédulas que estan en la base de datos, 
para una mayor facilidad para las personas encargadas del comedor se trabajará con archivos '.xlsx'
"""
# Módulos

from os import makedirs,startfile   
from tkinter import messagebox
from tkinter.filedialog import askopenfilenames
import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime
from shutil import move


# Variables

PlanNacional = {} # acá se almacenarán a todas las personas que se encuentren en el documento 'PlanNacional.xlsx'
PlanUsado = {}
Becas = {} # acá se almacenarán a todas las personas que se encuentren en el documento 'Becados.xlsx'
BecasUsadas = {}
ventas = [] # acá se almacenarán solanente las personas que no estan en plan nacional ni en becados.
cedulas = [] # acá se almacenarán todas las cédulas que han sido registradas en el durante el programa este en ejecución.



# Programa...

''''------------------------------------------------------------------------------------------------------------------'''

def mkdir (): # crea una carpeta donde se guardarán los archivos.
    makedirs(r'C:\SistemaComedor', exist_ok=True)


''''------------------------------------------------------------------------------------------------------------------'''

def isnotEmpty(data_structure, file): # verifica si los datos están en el formato correcto.
    if (len(data_structure) > 0):
        return True
    else:
        messagebox.showwarning("Formato incorrecto", f"El archivo {file} está en un formato incorrecto.")
        startfile(fr"C:\SistemaComedor\{file}")
        return False



''''------------------------------------------------------------------------------------------------------------------'''

def VAPN(): # verifica los datos de plan nacional.
    global PlanNacional
    Archivo = 'PlanNacional.xlsx' # funciona para cuando el programa tire un error y saber en cual parte esta el error.
    datos = [] # esta varible es temporal para guardar los datos que se encuentren en el archivo
    try:
        df = pd.read_excel(r'C:\SistemaComedor\PlanNacional.xlsx') # revisa los datos del archivo
        for valores in df.itertuples(index= False, name= None):
            datos.append(valores)
        PlanNacional = dict(datos)

        for clave in PlanNacional.keys(): 
            # Verifica si en las cédulas no hay ni letras.
            if str(clave).isnumeric() == False:
                PlanNacional.clear()
                break

        if isnotEmpty(PlanNacional, Archivo) == True:
            return True
    except FileNotFoundError:
        return False

VAPN()


''''------------------------------------------------------------------------------------------------------------------'''

def VAB(): # verifica los datos de las personas becadas.
    global Becas
    Archivo = 'Becados.xlsx' # funciona para cuando el programa tire un error y saber en cual parte esta el error.
    datos = [] # esta varible es temporal para guardar los datos que se encuentren en el archivo
    try:
        df = pd.read_excel(r'C:\SistemaComedor\Becados.xlsx') # revisa los datos del archivo
        for valores in df.itertuples(index= False, name= None):
            datos.append(valores)
        Becas = dict(datos)

        for clave in Becas.keys(): 
            # Verifica si en las cédulas no hay ni letras.
            if str(clave).isnumeric() == False:
                Becas.clear()
                break

        if isnotEmpty(Becas, Archivo) == True:
            return True
    except FileNotFoundError:
        return False
VAB()



''''------------------------------------------------------------------------------------------------------------------'''


def GuardarRegistro():
    global PlanUsado, BecasUsadas, ventas
    makedirs(r'C:\SistemaComedor\reportes', exist_ok=True)
    while True:
        try:
            fecha = str(datetime.today().strftime('%m-%y'))
            hoy = datetime.today().strftime('%d/%m/%y')
            hoy = pd.to_datetime(hoy, dayfirst= True,)
            archivo = load_workbook(r'C:\SistemaComedor\reportes\Reporte '+ fecha+'.xlsx')
            ws = archivo.active
            i = ws.max_row 
            i += 1
            ws.protection.disable() # deshabilita la protección del archivo para editarlo
            for cedula, nombre in PlanUsado.items():
                ws[f'A{i}']= cedula
                ws[f'B{i}']= nombre
                ws[f'C{i}']= 'Plan Nacional'
                ws[f'D{i}']= hoy
                i += 1
            for cedulab, nombreb in BecasUsadas.items():
                ws[f'A{i}'] = cedulab
                ws[f'B{i}'] = nombreb
                ws[f'C{i}'] = 'Becados'
                ws[f'D{i}'] = hoy
                i+= 1

            for cedulasv in ventas:
                ws[f'A{i}'] = cedulasv
                ws[f'B{i}'] = 'Desconocido'
                ws[f'C{i}'] = 'Ventas'
                ws[f'D{i}'] = hoy
                i += 1
            ws.protection.enable() # rehabilita la protección del archivo para que no pueda ser modificado. 
            archivo.save(r'C:\SistemaComedor\reportes\Reporte '+ fecha+'.xlsx')
            break

        except FileNotFoundError:
            fecha = str(datetime.today().strftime('%m-%y'))
            wb = Workbook()
            ws = wb.active
            ws.title = 'Registro'
            ws.append({1:'CEDULA', 2:'NOMBRE', 3:'GRUPO', 4:'FECHA'})
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            wb.save(fr'C:\SistemaComedor\reportes\Reporte '+fecha+'.xlsx')
            continue
        except PermissionError:
            messagebox.showwarning('Reporte abierto',f"Por Favor cierre el archivo antes de continuar 'Reporte {fecha}.xlsx'")
            continue


''''------------------------------------------------------------------------------------------------------------------'''

def en_encasodeerror(): # en el caso de que haya un erro el programa ejecutará esta función
    if VAB() is True and VAPN() is False:
            Archivo = 'PlanNacional.xlsx'
    else:
        Archivo = "Becados.xlsx"

    messagebox.showerror('FileNotFoundError',f'Agregue el archivo {Archivo} al directorio C:\SistemaComedor')
    source = askopenfilenames(title = 'Mover',filetypes=(('Excel files', '.xlsx'), ('Excel files', '.xlsx')))
    destination = 'C:\SistemaComedor'
    for files in source:

        move(files, destination)
    messagebox.showinfo('Importante', 'Abra nuevamente el programa')
    exit()


''''------------------------------------------------------------------------------------------------------------------'''

def abrir_reportes():
    # abre todos los reportes que han sido seleccionados.
    source = askopenfilenames(initialdir= r"C:\SistemaComedor\Reportes")
    for files in source:
        startfile(files)


''''------------------------------------------------------------------------------------------------------------------'''

def abrir_basededatos():
    # abre los archivos que han sido seleccionados
    source = askopenfilenames(initialdir= r"C:\SistemaComedor")
    for files in source:
        startfile(files)
    question = messagebox.askyesno('Importante', "¿Realizo alguno cambio en los archivos?")
    if  question == True: # si los reportes fueron modificados cierra el programa
        messagebox.showinfo('Importante', 'va a tener que abrir el programa nuevamente')
        exit()
    else:
        pass


''''------------------------------------------------------------------------------------------------------------------'''
# by milfuegosxd