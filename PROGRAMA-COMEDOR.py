# Módulos


from tkinter import END, Entry, PhotoImage, Tk, Label, Button, Toplevel, messagebox; from tkinter.filedialog import askopenfilenames
from os import makedirs, startfile
import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime; import time
import shutil
from PIL import Image, ImageTk


# variables

PersonasDelComedor = {} # guardará los datos de las personas que estan en el archivo
UsaronElComedor = {} # Datos de todas las personas que utilizaron el comedor.
CedulasRegistradas = [] # cedulas que fueron registradas durante la ejecución del programa.
actual = time.strftime('%d/%m/%y') # saca la fecha de hoy sin las horas.

# Funciones del programa.






def isnotEmpty(data_structure, file): 
    
    # Verifica si los archivos están en el formato correcto
    # si el archivo esta vacio es que no están en el formato correcto.
    
    if (len(data_structure) > 0):
        # si el archivo no esta vacio retorna true
        return True
    else:
        # si esta vacio muestra un mensaje
        # retorna false
        messagebox.showwarning("Formato incorrecto", f"El archivo {file} está en un formato incorrecto.")
        startfile(fr"C:\SistemaComedor\{file}")
        return False






def VERIFICACION_ARCHIVO(Archivo): 
    
    # La siguiente función verifica la integridad y el formato de los datos que se encuentran en 
    # en el archivo Comedor.xlsx

    datos = [] # Variable temporal y en esta se guardarán todos los datos que se encuentren en el archivo 
    info = [] # Esta variable es temportal y se guardarán los datos de nombre y sección de los archivos.



    try:
        
        # verifica si el archivo esta en la carpeta y también el formato de los datos.
        
        df = pd.read_excel(r'C:\SistemaComedor\Comedor.xlsx') 
        for values in df.itertuples(index= False, name= None):
            datos.append(values)

        for cedula, nombre, seccion, grupo in datos:

            info.append(nombre); info.append(seccion); info.append(grupo)
            PersonasDelComedor[cedula] = tuple(info)
            info.clear()
        
        datos.clear() # se borran todos los datos que estában guardados.
        
        for clave in PersonasDelComedor.keys():
            
            # Verifica si el lugar donde están deberían estar los números de cédulas
            # esta en el formato correcto.

            if str(clave).isnumeric() == False:
                PersonasDelComedor.clear()
                break
        if isnotEmpty(PersonasDelComedor, Archivo) == True:
            return True

    except FileNotFoundError:
        # Si el archivo no está devolverá False
        messagebox.showerror('ERROR ARCHIVO NO ENCONTRADO', f'PRESIONE ABRIR PARA MOVER EL ARCHIVO {Archivo} al diretorio C:\SistemaComedor')
        source = askopenfilenames(title = 'Mover',filetypes=(('Excel files', '.xlsx'), ('All files', '.')))
        destination = 'C:\SistemaComedor'
        for files in source:
            try:
                shutil.move(files, destination)
            except shutil.Error:
                pass
        return False

    except ValueError:
        # Si faltan datos y o están mal acomodados devolverá false
        messagebox.showwarning('FORMATO INCORRECTO', 'EL ARCHIVO DEBE IR EN EL SIGUIENTE FORMATO\nCedula, Nombre, Sección, Grupo')
        startfile(r'C:\SistemaComedor\Comedor.xlsx')
        return False






def GuardarRegistro():

    global UsaronElComedor
    fecha = str(datetime.today().strftime('%m-%y')) # Saca el mes y el año en el cual estamos, para utilizarlo en el nombre del reporte.
    hoy = datetime.today().strftime('%d/%m/%y %H:%M:%S') # Saca el día actual
    hoy = pd.to_datetime(hoy, dayfirst= True) # convirte la cadena de la variable hoy a el formato de fecha excel
    
    makedirs(r'C:\SistemaComedor\reportes', exist_ok=True) # crea el directorio donde se guardarán los reportes.
    
    while True:
        # El ciclo es para que continue mostrando el que en caso de error siga mostrando 
        # y pueda guardar el último dato que se ingreso antes de que ocurriese el error.
        try:
            archivo = load_workbook(r'C:\SistemaComedor\reportes\Reporte '+ fecha+'.xlsx')
            ws = archivo.active # worksheet 
            i = ws.max_row; i += 1 # busca la ultima celda sin datos.


            for cedula, datos in UsaronElComedor.items():

                ws[f'A{i}'] = cedula
                ws[f'B{i}'] = datos[0]
                ws[f'C{i}'] = datos[1]
                ws[f'D{i}'] = datos[2]
                ws[f'E{i}'] = hoy
                i += 1


            archivo.save(r'C:\SistemaComedor\reportes\Reporte '+ fecha+'.xlsx')
            break


        except FileNotFoundError:


            wb = Workbook()
            ws = wb.active
            ws.title = 'Registro'
            ws.append({1:'\tCEDULA', 2:'\tNOMBRE', 3: "\tSECCIÓN", 4:'\tGRUPO', 5:'\tFECHA'})
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            wb.save(fr'C:\SistemaComedor\reportes\Reporte '+fecha+'.xlsx')
            continue


        except PermissionError:
            messagebox.showwarning('Reporte abierto',f"Por Favor cierre el archivo antes de continuar 'Reporte {fecha}.xlsx'")
            continue






def abrir_reportes():
    dir = r'C:\SistemaComedor\Reportes'
    source = askopenfilenames(initialdir= dir)
    for files in source:
        startfile(files)






def centrar(ventana):
    # esta función centra la ventana dependiendo en que pantalla se trabaje
    ventana.geometry('1280x720')
    ventana.update_idletasks()
    w = ventana.winfo_width()
    h = ventana.winfo_height()
    extraW = ventana.winfo_screenwidth()-w
    extraH = ventana.winfo_screenheight()-h
    ventana.geometry("%dx%d%+d%+d" % (w,h,extraW/2,extraH/2))

def next(ventana):
    ventana.withdraw()






def back(cerrar, abrir):
    insertarCedulas()
    cerrar.destroy()
    abrir.deiconify()






def Second_win(selection):
    # dependiendo el boton que se presione abrirará una nueva ventana
    global entry, insertarCedulas







    def ventana(img):

        entry.delete(0,END)
        color = Toplevel(top)
        color.resizable(False, False)
        centrar(color)
        image = Image.open(img)
        tk_image = ImageTk.PhotoImage(image)
        Label(color, image = tk_image ).pack()
        def ocultar():
            color.destroy()
        color.after(2000,ocultar)
        color.mainloop()






    def Enter(event):
        insertarCedulas()






    def insertarCedulas():
        
        try:
            NumeroDeCedula = int(entry.get())

            # En esta parte se verificará si la persona ya a usado el comedor.

            # En el caso de que la persona no haya comido
            if NumeroDeCedula not in CedulasRegistradas:

                if NumeroDeCedula in PersonasDelComedor:
                    # si la persona esta en el archivo 

                    UsaronElComedor[NumeroDeCedula] = PersonasDelComedor[NumeroDeCedula]
                    ventana('green.png')
                    CedulasRegistradas.append(NumeroDeCedula)
                else:
                    # si la persona no esta en el archivo
                    ventana('gold.png')
                    UsaronElComedor[NumeroDeCedula] = ('Estudiante Regular', 'Sin especificar', 'Ventas')
            
            else:
                # se aparecerá una ventana emergente diciendo el que la persona ya usado el comedor
                ventana('red.png')

            GuardarRegistro()
            UsaronElComedor.clear()



        except ValueError:
            messagebox.showinfo('Valores no admitidos', 'Las cédulas no llevan letras.')
            entry.delete(0, END)






    next(root)
    if selection == 'InsertarCédulas' and VERIFICACION_ARCHIVO('Comedor.xlsx') == True:
        top = Toplevel(root)
        top.resizable(False, False)
        centrar(top)
        image = Image.open('INS.png')
        tk_image = ImageTk.PhotoImage(image)
        Label(top, image = tk_image ).pack()
        entry = Entry(top, justify='center', font = ('Comic Sans Ms', 50),width= 21, borderwidth=0)
        entry.bind('<Return>', Enter)
        entry.place(x = 219, y = 323)

    elif selection == 'AbrirBaseDeDatos':
        top = Toplevel(root)
        centrar(top)
    
    top.protocol('WM_DELETE_WINDOW', lambda: back(top, root))
    top.mainloop()






Font_tuple = ("Comic Sans MS", 30) 

makedirs(r'C:\SistemaComedor', exist_ok=True) # Crea la carpeta donde se guardarán los reportes

# INTERFAZ GRÁFICA.

root = Tk()
root.title('Comedor')
centrar(root)
root.wm_iconbitmap(True,'icono.ico')
root.resizable(False, False)

imagen = PhotoImage(file = 'COMEDOR.png')
Label(root, image= imagen, bd = 0).pack()

Bases = Button(root)
Bases.config(text= 'Revisar\narchivos', 
            font = Font_tuple,     
            width= 15, background='black', 
            activebackground='black', 
            foreground='white', 
            activeforeground='white', 
            borderwidth= 6,
            command=lambda: Second_win('AbrirBaseDeDatos')
            )
Bases.place(x = 215, y = 300)

BCedulas = Button(root)
BCedulas.config(text= 'Ingresar\ncédulas', 
                font = Font_tuple,
                width= 15, background='black', 
                activebackground='black', 
                foreground='white', 
                activeforeground='white', 
                borderwidth= 6, 
                command=lambda: Second_win('InsertarCédulas')
                )
BCedulas.place(x = 690, y = 300)

root.mainloop()